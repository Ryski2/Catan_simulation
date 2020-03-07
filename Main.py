from Simulation import Simulation
from Player import Strategies
import Globals as G
import numpy as np
import random
import sys
import time
import datetime
import os
from os import path
from multiprocessing import Pool
from openpyxl import Workbook
from openpyxl.styles import Alignment

sims = 100
if sims > 1:
    G.print_level = 0
else:
    G.print_level = 5

bar_len = 50

def simulation(args):
    (strats, board_str, random_order) = args
    seed = random.randrange(sys.maxsize)
    random.seed(seed)
    #print("Seed: " + str(seed))
    sim = Simulation(strats, board_str, random_order)
    turns, winners, points, roads, settlements, cities = sim.run()
    return (turns, winners, points, roads, settlements, cities)

if __name__ == "__main__":
    if not path.exists("Data"):
        os.mkdir("Data")

    if os.path.exists("Data/summary.csv") or os.path.exists("Data/turns.csv") \
            or os.path.exists("Data/wins.csv") or os.path.exists("Data/points.csv") \
            or os.path.exists("Data/roads.csv") or os.path.exists("Data/settlements.csv") \
            or os.path.exists("Data/cities.csv") or os.path.exists("Data/output.txt"):
        print("Make sure to rename the saved files before continuing")
        while True:
            text = input("Type y to continue anyways, n to exit: ")
            if text == "y":
                break
            if text == "n":
                sys.exit("Quitting.")

    random_order_list = [False] + [True] * 26

    strats_list = []
    board_str_list = []

    #1
    strats_list.append([{None}] * 4)
    board_str_list.append("basic")

    #2
    strats_list.append([{None}] * 4)
    board_str_list.append("basic")

    #3
    strats_list.append([{None}] * 4)
    board_str_list.append("random")

    #4
    strats_list.append([{None}] * 4)
    board_str_list.append("res_const")

    #5
    strats_list.append([{None}] * 4)
    board_str_list.append("val_const")

    #6
    strats_list.append([{Strategies.Trade}] * 4)
    board_str_list.append("random")

    #7
    strats_list.append([{Strategies.Trade}] + [{None}] * 3)
    board_str_list.append("random")

    #8
    strats_list.append([{Strategies.Road_Settlement_Ratio : 3}] * 4)
    board_str_list.append("random")

    #9
    strats_list.append([{Strategies.Road_Settlement_Ratio : 1}, {Strategies.Road_Settlement_Ratio : 2}, \
            {Strategies.Road_Settlement_Ratio : 3}, {Strategies.Road_Settlement_Ratio : 4}])
    board_str_list.append("random")

    #10
    strats_list.append([{Strategies.Road_Settlement_Ratio : 3}, {Strategies.Road_Settlement_Ratio : 4}, \
            {Strategies.Road_Settlement_Ratio : 5}, {Strategies.Road_Settlement_Ratio : 6}])
    board_str_list.append("random")

    #11
    strats_list.append([{Strategies.Prioritize_Settlements}] * 4)
    board_str_list.append("random")

    #12
    strats_list.append([{Strategies.Prioritize_Settlements}] * 2 + [{Strategies.Road_Settlement_Ratio : 4}] * 2)
    board_str_list.append("random")

    #13
    strats_list.append([{Strategies.Trade, Strategies.Prioritize_Settlements}] * 2 + [{Strategies.Trade : None, Strategies.Road_Settlement_Ratio : 4}])
    board_str_list.append("random")

    #14
    strats_list.append([{Strategies.Build_All}] * 4)
    board_str_list.append("random")

    #15
    strats_list.append([{Strategies.Build_All, Strategies.Prioritize_Settlements}] * 4)
    board_str_list.append("random")

    #16
    strats_list.append([{Strategies.Trade, Strategies.Prioritize_Settlements, Strategies.Build_All}, \
        {Strategies.Prioritize_Settlements, Strategies.Build_All}, {Strategies.Build_All}, {None}])
    board_str_list.append("random")

    #17
    strats_list.append([{Strategies.Avoid_Shore_and_Desert}] * 4)
    board_str_list.append("random")

    #18
    strats_list.append([{Strategies.Build_All, Strategies.Avoid_Shore_and_Desert}] * 4)
    board_str_list.append("random")

    #19
    strats_list.append([{Strategies.Trade, Strategies.Build_All, Strategies.Avoid_Shore_and_Desert}] * 4)
    board_str_list.append("random")

    #20
    strats_list.append([{Strategies.Trade, Strategies.Avoid_Shore_and_Desert}] * 4)
    board_str_list.append("random")

    #21
    strats_list.append([{Strategies.Robber_To_Opponent}] * 2 + [{None}] * 2)
    board_str_list.append("random")

    #22
    strats_list.append([{Strategies.Steal_From_Most_Resources}] * 2 + [{None}] * 2)
    board_str_list.append("random")

    #23
    strats_list.append([{Strategies.Discard_Most_Abundant}] * 2 + [{None}] * 2)
    board_str_list.append("random")

    #24
    strats_list.append([{Strategies.Robber_To_Opponent, Strategies.Steal_From_Most_Resources, Strategies.Discard_Most_Abundant}] * 2 + [{None}] * 2)
    board_str_list.append("random")

    #25
    strats_list.append([{Strategies.Robber_To_Opponent, Strategies.Steal_From_Most_Resources, Strategies.Discard_Most_Abundant}] * 4)
    board_str_list.append("random")

    #26
    strats_list.append([{Strategies.Trade, Strategies.Avoid_Shore_and_Desert, Strategies.Robber_To_Opponent, Strategies.Steal_From_Most_Resources, Strategies.Discard_Most_Abundant}] * 4)
    board_str_list.append("random")

    for i in range(1, 27):
        if i < 6:
            continue
        # Params
        strats = strats_list[i - 1]
        board_str = board_str_list[i - 1]
        random_order = random_order_list[i -1]

        if i == 1:
            sims = 100000
        else:
            sims = 10000

        total_turns = np.zeros((sims, 1), dtype=np.uint8)
        total_wins = np.zeros((sims, 4))
        total_points = np.zeros((sims, 4), dtype=np.uint8)
        total_roads = np.zeros((sims, 4), dtype=np.uint8)
        total_settlements = np.zeros((sims, 4), dtype=np.uint8)
        total_cities = np.zeros((sims, 4), dtype=np.uint8)

        print(sims)
        path_str = "Data/Section " + str(i)+ "/"
        if not path.exists(path_str):
            os.mkdir(path_str)

        start = time.perf_counter()
        p = Pool()
        for i, item in enumerate(p.imap_unordered(simulation, ((strats, board_str, random_order) for i in range(sims))), 0):
            if (i % 5 == 0):
                equals = round(i / sims * bar_len)
                spaces = bar_len - equals
                print("\r[" + "=" * equals + " " * spaces + "]\t" + str(i) + "/" + str(sims) + "\t" + "{:.2f}".format(i * 100 / sims)+ "% Done\t\t", end="")
            total_turns[i] = item[0]
            for winner in item[1]:
                total_wins[i, winner] = 1
            #total_wins[i, item[1]] = 1
            total_points[i] = item[2]
            total_roads[i] = item[3]
            total_settlements[i] = item[4]
            total_cities[i] = item[5]
        print("\r[" + "=" * bar_len + "]\t" + str(sims) + "/" + str(sims) + "\t" + "100% Done\t\t\t")
        p.close()
        p.terminate()
        p.join()

        turns = np.mean(total_turns, 0)
        wins = np.sum(total_wins, 0)
        points = np.mean(total_points, 0)
        roads = np.mean(total_roads, 0)
        settlements = np.mean(total_settlements, 0)
        cities = np.mean(total_cities, 0)


        end = time.perf_counter()
        output = ""
        output += "Turns:\n"
        output += "\tMean: " + str(turns) + "\n"
        output += "\tSt. Dev.: " + str(np.std(total_turns, 0)) + "\n"
        output += "Wins\n"
        output += "\tCount: " + str(wins) + "\n"
        output += "\tPercent: " + str(np.sum(total_wins, 0) / sims) + "\n"
        output += "Points:" + "\n"
        output += "\tMean: " + str(points) + "\n"
        output += "\tSt. Dev.: " + str(np.std(total_points, 0)) + "\n"
        output += "Roads:" + "\n"
        output += "\tMean: " + str(roads) + "\n"
        output += "\tSt. Dev.: " + str(np.std(total_roads, 0)) + "\n"
        output += "Settlements:" + "\n"
        output += "\tMean: " + str(settlements) + "\n"
        output += "\tSt. Dev.: " + str(np.std(total_settlements, 0)) + "\n"
        output += "Cities:" + "\n"
        output += "\tMean: " + str(cities) + "\n"
        output += "\tSt. Dev.: " + str(np.std(total_cities, 0)) + "\n"

        output += "Strategies Used: " + str(strats) + "\n"
        output += "Random Player Order: " + str(random_order) + "\n"
        output += "Total Elapsed: " + str(datetime.timedelta(seconds=end - start)) + "\n"

        summary = np.vstack((wins, points, roads, settlements, cities))
        np.savetxt(path_str + "summary.csv", np.vstack((np.pad(turns, (0, 3), mode='constant'), summary)), fmt="%f", delimiter=",")

        print(output)
        with open(path_str + "output.txt", "w") as f:
            print(output, file=f)

        wb = Workbook()
        ws = wb.active
        ws['A1'] = "Player"
        ws['B1'] = "Strategies"
        ws['C1'] = "Win Count"
        ws['D1'] = "Average Points"
        ws['E1'] = "Average Roads"
        ws['F1'] = "Average Settlements"
        ws['G1'] = "Average Cities"
        ws['H1'] = "Average Turn Count"
        ws.merge_cells('H2:H5')
        ws['H2'].alignment = Alignment(horizontal='center')
        ws['H2'] = turns[0]
        
        ws['A2'] = "Player 1"
        ws['A3'] = "Player 2"
        ws['A4'] = "Player 3"
        ws['A5'] = "Player 4"

        data_cells = ws['C2':'G5']
        for i, row in enumerate(data_cells):
            for j, cell in enumerate(row):
                if j == 0:
                    cell.value = int(np.transpose(summary)[i, j])
                else:
                    cell.value = float(np.transpose(summary)[i, j])   


        wb.save(path_str + "SummaryChart.xlsx")

        np.savetxt(path_str +"turns.csv", total_turns, fmt="%d", delimiter=",")
        np.savetxt(path_str + "wins.csv", total_turns, fmt="%d", delimiter=",")
        np.savetxt(path_str + "points.csv", total_points, fmt="%d", delimiter=",")
        np.savetxt(path_str + "roads.csv", total_roads, fmt="%d", delimiter=",")
        np.savetxt(path_str + "settlements.csv", total_settlements, fmt="%d", delimiter=",")
        np.savetxt(path_str + "cities.csv", total_cities, fmt="%d", delimiter=",")


    #total_turns = np.loadtxt("Data/turns.csv", dtype=np.unit8, delimiter=",")
    #total_points = np.loadtxt("Data/points.csv", dtype=np.unit8, delimiter=",")